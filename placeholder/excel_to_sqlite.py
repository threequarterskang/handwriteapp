from pathlib import Path
import sqlite3
import pandas as pd
from openpyxl import load_workbook


def load_named_excel_table(workbook_path: Path, object_name: str) -> pd.DataFrame:
    workbook = load_workbook(workbook_path, data_only=True)

    for worksheet in workbook.worksheets:
        if object_name in worksheet.tables:
            table_range = worksheet.tables[object_name].ref
            rows = list(worksheet[table_range])
            headers = [cell.value for cell in rows[0]]
            records = [[cell.value for cell in row] for row in rows[1:]]
            return pd.DataFrame(records, columns=headers)

    if object_name in workbook.sheetnames:
        return pd.read_excel(workbook_path, sheet_name=object_name)

    available_sheets = ", ".join(workbook.sheetnames)
    raise ValueError(
        f"Could not find an Excel table or worksheet named {object_name!r}. "
        f"Available worksheets: {available_sheets}"
    )


def normalize_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    normalized = dataframe.copy()
    normalized.columns = [str(column).strip() for column in normalized.columns]
    normalized = normalized.dropna(how="all")
    normalized = normalized.where(pd.notna(normalized), None)
    return normalized


def format_date_columns(
    dataframe: pd.DataFrame,
    date_columns: list[str]
) -> pd.DataFrame:
    """
    将指定列格式化为 YYYY.MM.DD
    """
    df = dataframe.copy()
    for col in date_columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%Y.%m.%d")
    return df


def import_excel_with_partial_schema(
    workbook_path: Path,
    database_path: Path,
    excel_object_name: str,
    sqlite_table_name: str,
    column_types: dict[str, str],
    date_columns: list[str] | None = None
) -> tuple[int, int]:

    dataframe = load_named_excel_table(workbook_path, excel_object_name)
    dataframe = normalize_dataframe(dataframe)

    if date_columns:
        dataframe = format_date_columns(dataframe, date_columns)

    with sqlite3.connect(database_path) as connection:
        cursor = connection.cursor()

        columns_sql = []
        for col in dataframe.columns:
            if col in column_types:
                columns_sql.append(f'"{col}" {column_types[col]}')
            else:
                columns_sql.append(f'"{col}" TEXT')

        cursor.executescript(f"""
            DROP TABLE IF EXISTS {sqlite_table_name};
            CREATE TABLE {sqlite_table_name} (
                {', '.join(columns_sql)}
            );
        """)

        dataframe.to_sql(
            sqlite_table_name,
            connection,
            if_exists="append",
            index=False
        )

    return len(dataframe.index), len(dataframe.columns)